"""Tests for scripts/load_leadsun_edge_accounts.py"""

import json
import os

import openpyxl
import pytest
from cryptography.fernet import Fernet

from scripts.load_leadsun_edge_accounts import (
    _UPSERT_ACCOUNT_SQL,
    load_leadsun_edge_accounts,
    load_local_settings_into_env,
    read_accounts_from_xlsx,
    refuse_if_prod,
)


def _write_xlsx(tmp_path, rows, headers=("Username", "Password"), filename="accounts.xlsx"):
    """Writes a minimal .xlsx with the given header row and data rows."""
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(list(headers))
    for row in rows:
        sheet.append(list(row))
    path = tmp_path / filename
    wb.save(path)
    return path


class TestRefuseIfProd:
    def test_raises_for_prod(self):
        with pytest.raises(SystemExit):
            refuse_if_prod("Prod")

    def test_does_not_raise_for_dev(self):
        refuse_if_prod("Dev")  # must not raise


class TestLoadLocalSettingsIntoEnv:
    def test_returns_false_when_file_missing(self, tmp_path):
        assert load_local_settings_into_env(project_root=tmp_path) is False

    def test_loads_values_into_environment(self, tmp_path, monkeypatch):
        monkeypatch.delenv("SOME_TEST_KEY", raising=False)
        (tmp_path / "local.settings.json").write_text(
            json.dumps({"Values": {"SOME_TEST_KEY": "some-value"}})
        )
        assert load_local_settings_into_env(project_root=tmp_path) is True
        assert os.environ["SOME_TEST_KEY"] == "some-value"


class TestReadAccountsFromXlsx:
    def test_reads_username_and_password_pairs(self, tmp_path):
        path = _write_xlsx(tmp_path, [("alice", "pw1"), ("bob", "pw2")])

        accounts = read_accounts_from_xlsx(path)

        assert accounts == [("alice", "pw1"), ("bob", "pw2")]

    def test_locates_columns_by_header_name_not_position(self, tmp_path):
        """Column order in the source file shouldn't matter -- located
        by header text, not fixed position."""
        path = _write_xlsx(tmp_path, [("pw1", "alice")], headers=("Password", "Username"))

        accounts = read_accounts_from_xlsx(path)

        assert accounts == [("alice", "pw1")]

    def test_missing_username_header_raises_clearly(self, tmp_path):
        path = _write_xlsx(tmp_path, [("alice", "pw1")], headers=("Name", "Password"))

        with pytest.raises(ValueError, match="Username.*Password"):
            read_accounts_from_xlsx(path)

    def test_missing_password_header_raises_clearly(self, tmp_path):
        path = _write_xlsx(tmp_path, [("alice", "pw1")], headers=("Username", "Secret"))

        with pytest.raises(ValueError, match="Username.*Password"):
            read_accounts_from_xlsx(path)

    def test_empty_sheet_raises_clearly(self, tmp_path):
        wb = openpyxl.Workbook()
        path = tmp_path / "empty.xlsx"
        wb.save(path)

        with pytest.raises(ValueError, match="empty"):
            read_accounts_from_xlsx(path)

    def test_row_with_missing_username_is_skipped_not_fatal(self, tmp_path, caplog):
        path = _write_xlsx(tmp_path, [(None, "pw1"), ("bob", "pw2")])

        with caplog.at_level("WARNING"):
            accounts = read_accounts_from_xlsx(path)

        assert accounts == [("bob", "pw2")]
        assert any("skipping" in rec.message.lower() for rec in caplog.records)

    def test_row_with_missing_password_is_skipped_not_fatal(self, tmp_path):
        path = _write_xlsx(tmp_path, [("alice", None), ("bob", "pw2")])

        accounts = read_accounts_from_xlsx(path)

        assert accounts == [("bob", "pw2")]

    def test_skipped_row_does_not_log_the_actual_password(self, tmp_path, caplog):
        """A row is only skipped when username OR password is missing --
        but if password IS present (only username is missing), it must
        not appear in plaintext in the log output."""
        path = _write_xlsx(tmp_path, [(None, "supersecretpassword123")])

        with caplog.at_level("WARNING"):
            read_accounts_from_xlsx(path)

        full_log_text = "\n".join(rec.message for rec in caplog.records)
        assert "supersecretpassword123" not in full_log_text

    def test_reads_only_the_first_sheet(self, tmp_path):
        wb = openpyxl.Workbook()
        wb.active.append(["Username", "Password"])
        wb.active.append(["alice", "pw1"])
        other_sheet = wb.create_sheet("Sheet2")
        other_sheet.append(["Username", "Password"])
        other_sheet.append(["should-not-appear", "pw2"])
        path = tmp_path / "multi.xlsx"
        wb.save(path)

        accounts = read_accounts_from_xlsx(path)

        assert accounts == [("alice", "pw1")]

    def test_usernames_are_stripped_of_whitespace(self, tmp_path):
        path = _write_xlsx(tmp_path, [("  alice  ", "pw1")])

        accounts = read_accounts_from_xlsx(path)

        assert accounts == [("alice", "pw1")]


class TestLoadLeadsunEdgeAccounts:
    @pytest.fixture(autouse=True)
    def _real_key(self, monkeypatch):
        monkeypatch.setenv("LEADSUN_EDGE_ACCOUNTS_ENCRYPTION_KEY", Fernet.generate_key().decode())

    def test_upserts_each_account_with_an_encrypted_password(
        self, tmp_path, mocker, mock_conn, mock_cursor
    ):
        mocker.patch("shared.sql_client.get_connection", return_value=mock_conn)
        path = _write_xlsx(tmp_path, [("alice", "pw1"), ("bob", "pw2")])

        count = load_leadsun_edge_accounts(path)

        assert count == 2
        merge_calls = [c for c in mock_cursor.execute.call_args_list if c.args[0] == _UPSERT_ACCOUNT_SQL]
        assert len(merge_calls) == 2

        usernames_bound = {c.args[1] for c in merge_calls}
        assert usernames_bound == {"alice", "bob"}

        # The SAME encrypted value is bound twice per call (once for the
        # UPDATE branch, once for the INSERT branch) -- confirm both are
        # identical within a single call, and neither is the plaintext.
        for call in merge_calls:
            username, encrypted_1, encrypted_2 = call.args[1], call.args[2], call.args[3]
            assert encrypted_1 == encrypted_2
            assert encrypted_1 not in ("pw1", "pw2")

    def test_commits_once_after_all_accounts_are_upserted(
        self, tmp_path, mocker, mock_conn, mock_cursor
    ):
        mocker.patch("shared.sql_client.get_connection", return_value=mock_conn)
        path = _write_xlsx(tmp_path, [("alice", "pw1"), ("bob", "pw2"), ("carol", "pw3")])

        load_leadsun_edge_accounts(path)

        mock_conn.commit.assert_called_once()

    def test_rolls_back_and_reraises_on_failure_without_partial_commit(
        self, tmp_path, mocker, mock_conn, mock_cursor
    ):
        mocker.patch("shared.sql_client.get_connection", return_value=mock_conn)
        mock_cursor.execute.side_effect = RuntimeError("connection lost")
        path = _write_xlsx(tmp_path, [("alice", "pw1")])

        with pytest.raises(RuntimeError, match="connection lost"):
            load_leadsun_edge_accounts(path)

        mock_conn.rollback.assert_called_once()
        mock_conn.commit.assert_not_called()

    def test_closes_cursor_and_connection_even_on_failure(
        self, tmp_path, mocker, mock_conn, mock_cursor
    ):
        mocker.patch("shared.sql_client.get_connection", return_value=mock_conn)
        mock_cursor.execute.side_effect = RuntimeError("connection lost")
        path = _write_xlsx(tmp_path, [("alice", "pw1")])

        with pytest.raises(RuntimeError):
            load_leadsun_edge_accounts(path)

        mock_cursor.close.assert_called_once()
        mock_conn.close.assert_called_once()

    def test_each_account_gets_a_genuinely_different_ciphertext(
        self, tmp_path, mocker, mock_conn, mock_cursor
    ):
        """Even two accounts sharing the exact same real password must
        not end up with the same EncryptedPassword value -- confirms the
        loader doesn't accidentally reuse/cache a single encryption
        result across rows."""
        mocker.patch("shared.sql_client.get_connection", return_value=mock_conn)
        path = _write_xlsx(tmp_path, [("alice", "sharedpassword"), ("bob", "sharedpassword")])

        load_leadsun_edge_accounts(path)

        merge_calls = [c for c in mock_cursor.execute.call_args_list if c.args[0] == _UPSERT_ACCOUNT_SQL]
        encrypted_values = [c.args[2] for c in merge_calls]
        assert encrypted_values[0] != encrypted_values[1]

    def test_raises_clearly_when_encryption_key_missing(
        self, tmp_path, mocker, mock_conn, mock_cursor, monkeypatch
    ):
        monkeypatch.delenv("LEADSUN_EDGE_ACCOUNTS_ENCRYPTION_KEY", raising=False)
        mocker.patch("shared.sql_client.get_connection", return_value=mock_conn)
        path = _write_xlsx(tmp_path, [("alice", "pw1")])

        from shared.encryption_utils import EncryptionConfigError

        with pytest.raises(EncryptionConfigError):
            load_leadsun_edge_accounts(path)


class TestUpsertSqlIsAMergeNotAPlainInsert:
    def test_is_a_merge(self):
        """Idempotent/re-runnable, consistent with every other loader in
        this project using upsert semantics rather than a one-shot
        insert that would fail (or duplicate) on re-run."""
        assert "MERGE LeadsunEdgeAccounts AS target" in _UPSERT_ACCOUNT_SQL
        assert "ON target.Username = source.Username" in _UPSERT_ACCOUNT_SQL

    def test_updates_the_updated_at_timestamp_on_an_existing_row(self):
        assert "UpdatedAt = SYSDATETIMEOFFSET()" in _UPSERT_ACCOUNT_SQL

    def test_does_not_touch_loaded_at_on_an_existing_row(self):
        """LoadedAt should only ever be set once, at initial INSERT --
        the UPDATE branch must not also refresh it."""
        update_branch = _UPSERT_ACCOUNT_SQL.split("WHEN MATCHED THEN UPDATE SET")[1].split(
            "WHEN NOT MATCHED"
        )[0]
        assert "LoadedAt" not in update_branch
